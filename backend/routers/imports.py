"""Excel import — dry-run + commit for legacy breakdown workbook."""
import hashlib
from datetime import datetime, timezone
from io import BytesIO
from typing import List, Dict, Any, Optional
from fastapi import APIRouter, Depends, UploadFile, File, HTTPException
from db import get_db
from deps import require_admin, write_audit
from models import uid, now_utc

router = APIRouter(prefix="/api/imports", tags=["imports"])


def _norm(s):
    return (str(s or "").strip().lower())


def _parse_date(v):
    """Return YYYY-MM-DD from many possible Excel forms."""
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.date().isoformat()
    # excel dates come as datetime from openpyxl; string fallback
    try:
        return datetime.fromisoformat(str(v).split(" ")[0]).date().isoformat()
    except Exception:
        for fmt in ("%d-%m-%Y", "%d/%m/%Y", "%Y-%m-%d", "%m/%d/%Y", "%d-%b-%Y"):
            try:
                return datetime.strptime(str(v).strip(), fmt).date().isoformat()
            except Exception:
                continue
    return None


def _parse_time_to_iso(date_str: Optional[str], time_val):
    if not date_str or time_val is None or time_val == "":
        return None
    if isinstance(time_val, datetime):
        return time_val.replace(tzinfo=timezone.utc).isoformat()
    # try HH:MM or HH:MM:SS
    for fmt in ("%H:%M:%S", "%H:%M", "%I:%M %p", "%I:%M:%S %p"):
        try:
            t = datetime.strptime(str(time_val).strip(), fmt).time()
            return datetime.combine(datetime.fromisoformat(date_str).date(), t).replace(tzinfo=timezone.utc).isoformat()
        except Exception:
            continue
    return None


def _classify_type(txt: str) -> str:
    t = _norm(txt)
    if not t:
        return "other"
    if "mech" in t:
        return "mechanical"
    if "elec" in t:
        return "electrical"
    if "proc" in t:
        return "process"
    if "instr" in t or "sensor" in t:
        return "instrumentation"
    if "util" in t or "steam" in t:
        return "utility"
    if "oper" in t:
        return "operator_error"
    if "plan" in t:
        return "planned"
    return "other"


async def _parse_rows(raw_bytes: bytes) -> List[Dict[str, Any]]:
    from openpyxl import load_workbook
    wb = load_workbook(BytesIO(raw_bytes), data_only=True)
    ws = wb.active
    # find header row
    header_row_idx = 1
    headers: List[str] = []
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=5, values_only=True), start=1):
        if row and sum(1 for c in row if c) >= 5:
            headers = [str(c or "").strip() for c in row]
            header_row_idx = i
            break
    if not headers:
        raise HTTPException(status_code=400, detail="Could not find header row")

    def col(name_options):
        for want in name_options:
            for idx, h in enumerate(headers):
                if _norm(h) == _norm(want):
                    return idx
        # partial match
        for want in name_options:
            for idx, h in enumerate(headers):
                if _norm(want) in _norm(h):
                    return idx
        return None

    idx_email = col(["email"])
    idx_area = col(["area of breakdown", "area"])
    idx_equip = col(["equipment", "select equipment pc21", "select equipment pc32",
                      "select equipment pc 36", "select equipment kkr", "select equipment twz",
                      "select equipment bcp", "select equipment"])
    idx_desc = col(["breakdown description", "description"])
    idx_type = col(["breakdown type", "type"])
    idx_date = col(["date of breakdown", "date"])
    idx_start = col(["breakdown start time", "start time"])
    idx_end = col(["breakdown end time", "end time", "completion time"])
    idx_action = col(["action taken"])
    idx_att = col(["attended by"])
    idx_status = col(["breakdown status", "status"])
    idx_spares = col(["spares used (sap code only)", "spares used", "spares"])

    rows = []
    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        if not row or not any(row):
            continue

        def get(i):
            return row[i] if (i is not None and i < len(row)) else None

        area = get(idx_area)
        equip = get(idx_equip) or area
        desc = get(idx_desc) or ""
        if not equip and not desc:
            continue
        d = _parse_date(get(idx_date))
        start_iso = _parse_time_to_iso(d, get(idx_start))
        end_iso = _parse_time_to_iso(d, get(idx_end))
        spares_raw = str(get(idx_spares) or "")
        spares = []
        if spares_raw.strip():
            for tok in spares_raw.replace(";", ",").replace("|", ",").split(","):
                t = tok.strip()
                if t:
                    spares.append({"sap_code": t, "qty": 1, "cost": 0.0})

        rows.append({
            "email": (get(idx_email) or "").strip() if get(idx_email) else None,
            "area": (str(area).strip() if area else None),
            "equipment": (str(equip).strip() if equip else None),
            "description": (str(desc).strip() if desc else ""),
            "breakdown_type": _classify_type(get(idx_type)),
            "date": d,
            "start_iso": start_iso,
            "end_iso": end_iso,
            "action_taken": (str(get(idx_action) or "").strip() or None),
            "attended_by": (str(get(idx_att) or "").strip() or None),
            "status_text": (str(get(idx_status) or "").strip() or "closed"),
            "spares": spares,
        })
    return rows


def _row_hash(r: dict) -> str:
    s = "|".join([r.get("email") or "", r.get("date") or "",
                    r.get("start_iso") or "", r.get("end_iso") or "",
                    r.get("equipment") or "", r.get("description") or ""])
    return hashlib.sha256(s.encode()).hexdigest()


async def _match_machine(db, equipment_text: str, area_text: str | None) -> Optional[dict]:
    text = _norm(f"{equipment_text} {area_text or ''}")
    # exact code / name match
    machines = await db.machines.find({}, {"_id": 0, "id": 1, "code": 1, "name": 1, "line_id": 1, "is_packing": 1}).to_list(2000)
    # Best fuzzy: contains name; also try tokens
    scored = []
    for m in machines:
        n = _norm(m["name"])
        c = _norm(m["code"])
        score = 0
        if n and n in text:
            score = max(score, 90)
        if c and c in text:
            score = max(score, 95)
        # per-token
        for tok in n.split():
            if len(tok) > 2 and tok in text:
                score = max(score, 70)
        if score > 0:
            scored.append((score, m))
    if not scored:
        return None
    scored.sort(key=lambda x: -x[0])
    return scored[0][1]


@router.post("/breakdowns/dry-run")
async def dry_run(file: UploadFile = File(...), admin=Depends(require_admin)):
    raw = await file.read()
    rows = await _parse_rows(raw)
    db = get_db()
    matched = 0
    unmatched: list[dict] = []
    dup_count = 0
    seen = set()
    for r in rows:
        h = _row_hash(r)
        if h in seen or await db.breakdowns.find_one({"import_hash": h}, {"_id": 0, "id": 1}):
            dup_count += 1
            continue
        seen.add(h)
        m = await _match_machine(db, r["equipment"] or "", r["area"])
        if m:
            matched += 1
        else:
            unmatched.append({"equipment": r["equipment"], "area": r["area"], "date": r["date"]})
    return {"ok": True, "data": {
        "total_rows": len(rows),
        "matched": matched,
        "unmatched": len(unmatched),
        "duplicates": dup_count,
        "unmatched_examples": unmatched[:15],
    }}


@router.post("/breakdowns/commit")
async def commit(file: UploadFile = File(...), admin=Depends(require_admin)):
    raw = await file.read()
    rows = await _parse_rows(raw)
    db = get_db()
    inserted = 0
    skipped = 0
    unmatched = 0
    for r in rows:
        h = _row_hash(r)
        if await db.breakdowns.find_one({"import_hash": h}, {"_id": 0, "id": 1}):
            skipped += 1
            continue
        m = await _match_machine(db, r["equipment"] or "", r["area"])
        if not m or m.get("is_packing"):
            unmatched += 1
            continue
        line = await db.production_lines.find_one({"id": m["line_id"]}, {"_id": 0})
        year = datetime.now(timezone.utc).year
        c = await db.counters.find_one_and_update(
            {"key": f"BD_{year}"}, {"$inc": {"seq": 1}}, upsert=True, return_document=True,
        )
        ticket = f"BD-{year}-{c['seq']:06d}"
        wo_c = await db.counters.find_one_and_update(
            {"key": f"WO_{year}"}, {"$inc": {"seq": 1}}, upsert=True, return_document=True,
        )
        wo_no = f"WO-{year}-{wo_c['seq']:06d}"

        start = r["start_iso"]
        end = r["end_iso"]
        dur = None
        if start and end:
            try:
                a = datetime.fromisoformat(start)
                b = datetime.fromisoformat(end)
                dur = max(0, int((b - a).total_seconds()))
            except Exception:
                dur = None
        status = "closed" if end else "open"
        now = now_utc().isoformat()

        bd = {
            "id": uid(), "ticket_no": ticket,
            "plant_id": line["plant_id"], "line_id": m["line_id"], "machine_id": m["id"],
            "reported_by": admin["id"], "reporter_email": r.get("email"),
            "area_text": r.get("area"), "equipment_text": r.get("equipment"),
            "description": r.get("description") or "",
            "failure_mode_id": None,
            "breakdown_type": r.get("breakdown_type") or "other",
            "date_of_breakdown": r.get("date") or (start[:10] if start else now[:10]),
            "breakdown_start_ts": start,
            "breakdown_end_ts": end,
            "duration_seconds": dur,
            "status": status,
            "severity": "medium",
            "photos": [], "work_order_id": None,
            "import_hash": h, "import_source": "excel",
            "created_at": now, "updated_at": now,
            "created_by": admin["id"], "updated_by": admin["id"],
        }

        wo = {
            "id": uid(), "wo_no": wo_no, "breakdown_id": bd["id"],
            "plant_id": line["plant_id"], "line_id": m["line_id"], "machine_id": m["id"],
            "type": "corrective", "priority": "p3",
            "status": "closed" if end else "open",
            "assigned_to": None,
            "assigned_at": start, "accepted_at": start,
            "repair_started_at": start, "repair_completed_at": end,
            "closed_at": end,
            "response_time_seconds": 0,
            "repair_time_seconds": dur,
            "close_time_seconds": 0,
            "action_taken": r.get("action_taken"), "root_cause": None,
            "spares_used": r.get("spares") or [],
            "assignment_history": [],
            "created_at": now, "updated_at": now,
            "created_by": admin["id"], "updated_by": admin["id"],
        }
        await db.breakdowns.insert_one(bd)
        await db.work_orders.insert_one(wo)
        await db.breakdowns.update_one({"id": bd["id"]}, {"$set": {"work_order_id": wo["id"]}})
        # timeline (backdated)
        if start:
            await db.timeline_events.insert_one({
                "id": uid(), "at": start, "plant_id": line["plant_id"], "line_id": m["line_id"],
                "machine_id": m["id"], "actor_id": admin["id"], "kind": "breakdown.created",
                "payload": {"ticket_no": ticket, "wo_no": wo_no, "imported": True},
                "source": "import", "ref_id": bd["id"],
                "created_at": now, "updated_at": now,
            })
        if end:
            await db.timeline_events.insert_one({
                "id": uid(), "at": end, "plant_id": line["plant_id"], "line_id": m["line_id"],
                "machine_id": m["id"], "actor_id": admin["id"], "kind": "breakdown.closed",
                "payload": {"ticket_no": ticket, "duration_seconds": dur, "imported": True},
                "source": "import", "ref_id": bd["id"],
                "created_at": now, "updated_at": now,
            })
        inserted += 1

    await write_audit(admin["id"], "import.commit", "import", "breakdowns",
                       after={"inserted": inserted, "skipped": skipped, "unmatched": unmatched})
    return {"ok": True, "data": {"inserted": inserted, "skipped": skipped, "unmatched": unmatched}}
